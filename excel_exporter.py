import io
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from comparator import BatchComparisonResult


class ExcelExporter:

    @staticmethod
    def generate(res: BatchComparisonResult) -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Impurity Matrix"

        font_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        font_subhead = Font(name="Calibri", size=10, bold=True)
        font_regular = Font(name="Calibri", size=10)

        fill_header = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        fill_sub = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
        fill_main = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")

        b_thin = Side(border_style="thin", color="000000")
        std_border = Border(left=b_thin, right=b_thin, top=b_thin, bottom=b_thin)

        # 3-Tier Header Structure matching your reference image
        # Row 1: Name of Impurity (Spans across impurity columns)
        # Row 2: RT values
        # Row 3: RRT values

        ws.cell(row=1, column=1, value="Sr. No.").alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=1, column=2, value="Batch No.").alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("C1:C3")
        ws.cell(row=1, column=3, value="Name of Impurity").alignment = Alignment(horizontal="center", vertical="center")

        num_meta_cols = 3
        start_col = num_meta_cols + 1

        for i, col in enumerate(res.master_columns):
            c_idx = start_col + i
            # Row 1: Impurity Name
            ws.cell(row=1, column=c_idx, value=col.peak_name)
            # Row 2: RT
            ws.cell(row=2, column=c_idx, value=col.rt)
            # Row 3: RRT
            ws.cell(row=3, column=c_idx, value=col.rrt)

        ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=start_col + len(res.master_columns) - 1)

        for r in range(1, 4):
            for c in range(1, start_col + len(res.master_columns)):
                cell = ws.cell(row=r, column=c)
                cell.font = font_header if r == 1 else font_subhead
                cell.fill = fill_header if r == 1 else fill_sub
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = std_border

        # Populate Data Rows
        data_start_row = 4
        for row_idx, b_row in enumerate(res.batch_rows, start=data_start_row):
            ws.cell(row=row_idx, column=1, value=b_row["Sr. No."]).alignment = Alignment(horizontal="center")
            ws.cell(row=row_idx, column=2, value=b_row["Batch No."]).alignment = Alignment(horizontal="left")
            ws.cell(row=row_idx, column=3, value=b_row["Injection Name"]).alignment = Alignment(horizontal="left")

            for col_i, col in enumerate(res.master_columns):
                c_idx = start_col + col_i
                val = b_row.get(col.rt, "")
                cell = ws.cell(row=row_idx, column=c_idx, value=val)
                cell.alignment = Alignment(horizontal="right")
                cell.font = font_regular
                cell.border = std_border
                if col.is_main_peak:
                    cell.fill = fill_main

            for c in range(1, 4):
                ws.cell(row=row_idx, column=c).border = std_border
                ws.cell(row=row_idx, column=c).font = font_regular

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
